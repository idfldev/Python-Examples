import openpyxl
import re
import requests







def check_cas_number(cas_number):
    """
    Function to check the validity of a CAS number.

    Args:
        cas_number (str): The CAS number to be checked.

    Returns:
        bool: True if the CAS number is valid, False otherwise.
    """
    pattern = r'^\d{2,7}-\d{2}-\d$'  # Regular expression pattern for CAS numbers
    if re.match(pattern, cas_number):
        return True
    else:
        return False












# Open the XLS file
workbook = openpyxl.load_workbook('example.xlsx')

# Get the two worksheets that contain the elements to compare
sheet1 = workbook['Sheet1']
sheet2 = workbook['Sheet2']

# Get the elements in column A of the first worksheet
elements1 = []
for row in sheet1.iter_rows(min_row=2, values_only=True):
    element = row[0]
    if element is not None:
        elements1.append(element.strip().lower())

# Get the elements in column A of the second worksheet
elements2 = []
for row in sheet2.iter_rows(min_row=2, values_only=True):
    element = row[0]
    if element is not None:
        elements2.append(element.strip().lower())

# Compare the two lists of elements
in_common = set(elements1) & set(elements2)
only_in_elements1 = set(elements1) - set(elements2)
only_in_elements2 = set(elements2) - set(elements1)

# Print the results
print("Elements in common:")
for element in sorted(in_common):
    print(element)
print()

print("Elements only in sheet 1:")
for element in sorted(only_in_elements1):
    print(element)
print()

print("Elements only in sheet 2:")
for element in sorted(only_in_elements2):
    print(element)
